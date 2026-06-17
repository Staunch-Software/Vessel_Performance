"""
==============================================================================
  WNI DATA MAPPER - AM_KIRTI
==============================================================================

  FILENAMES BEING SEARCHED:
  -------------------------
  Analysis File: AM_KIRTI_MARIAPPS_NEW.xlsx
  WNI File:      AM_KIRTI.xlsx

  HOW TO RUN:
  -----------
  Run from ROOT of your project (Data_ingestion_pipeline):
      python backend/excel_processing/wni_mapper.py

==============================================================================
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os

# ==============================================================================
#  SETTINGS & PATHS
# ==============================================================================

# --- LOCAL (original hardcoded paths — uncomment to use) ---
# INPUT_DIR  = r'C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\backend\excel_processing\data\new_reports'
# OUTPUT_DIR = r'C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\backend\excel_processing\output'
# --- VM / CROSS-PLATFORM (env var, defaults relative to this file) ---
_BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.getenv("WNI_MAPPER_INPUT_DIR",  os.path.join(_BASE_DIR, 'data', 'new_reports'))
OUTPUT_DIR = os.getenv("WNI_MAPPER_OUTPUT_DIR", os.path.join(_BASE_DIR, 'output'))


ANALYSIS_FILENAME = os.path.join(INPUT_DIR,  'GCL_TAPI_MARIAPPS.xlsx')
WNI_FILENAME      = os.path.join(INPUT_DIR,  'GCL_TAPI.xlsx')
OUTPUT_FILENAME   = os.path.join(OUTPUT_DIR, 'GCL_TAPI_NEWLY.xlsx')

# List of columns to attempt to fill from the WNI file
FILL_COLS = [
    'Loading_Cond',
    'STW_kn',
    'SOG_kn',
    'Distance_nm',
    'Shaft_Power_kW',
    'Shaft_RPM',
    'ME_FOC_MT',
    'Est_Power_kW',
    'SFOC_gkWh',
    'True_Wind_Dir_deg',
    'P_wind_kW',
    'P_wave_kW',
    'P_temp_kW',
    'VTI',
    'Power_Dev_pct',
    'Speed_Loss_pct',
]

# ==============================================================================
#  MAIN PROCESSING LOGIC
# ==============================================================================

def main():
    print("\n" + "="*60)
    print("  WNI DATA MAPPER - Starting")
    print("="*60)

    # 1. Check files exist
    if not os.path.exists(ANALYSIS_FILENAME):
        print(f"\n  ERROR: Analysis file not found: {ANALYSIS_FILENAME}")
        return
    if not os.path.exists(WNI_FILENAME):
        print(f"\n  ERROR: WNI file not found: {WNI_FILENAME}")
        return

    # 2. Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 3. Load the Excel files into DataFrames
    print(f"\n  Loading Analysis : {ANALYSIS_FILENAME}")
    try:
        analysis = pd.read_excel(ANALYSIS_FILENAME)
        print(f"    -> {len(analysis)} rows loaded")
    except Exception as e:
        print(f"  ERROR reading Analysis file: {e}")
        return

    print(f"  Loading WNI      : {WNI_FILENAME}")
    try:
        wni = pd.read_excel(WNI_FILENAME)
        print(f"    -> {len(wni)} rows loaded")
    except Exception as e:
        print(f"  ERROR reading WNI file: {e}")
        return

    # 4. Create Matching Keys (Date + Time_UTC)
    print("\n  Matching rows by Date + Time_UTC...")

    analysis['Date'] = pd.to_datetime(analysis['Date']).dt.date
    wni['Date']      = pd.to_datetime(wni['Date']).dt.date

    analysis['_key'] = analysis['Date'].astype(str) + '|' + analysis['Time_UTC'].astype(str).str.strip().str[:5]
    wni['_key']      = wni['Date'].astype(str) + '|' + wni['Time_UTC'].apply(lambda t: str(t).strip()[:5])

    # 5. Merge the data
    cols_to_pull = [c for c in FILL_COLS if c in wni.columns]
    wni_slim = wni[['_key'] + cols_to_pull].drop_duplicates('_key')

    merged = analysis.merge(wni_slim, on='_key', how='left', suffixes=('', '_wni'))

    matched_count = int(merged['_key'].isin(wni['_key']).sum())
    print(f"    -> {matched_count} rows matched | {len(analysis)-matched_count} rows unmatched")

    # 6. Fill missing values
    print("\n  Filling NULL and ZERO values from WNI...")
    filled_cells = []
    filled_total = 0

    for col in cols_to_pull:
        wni_col = col + '_wni'
        if wni_col not in merged.columns:
            continue

        try:
            mask = (
                (merged[col].isnull() | (merged[col] == 0)) &
                (merged[wni_col].notnull()) &
                (merged[wni_col] != 0)
            )
        except Exception:
            mask = (merged[col].isnull() & merged[wni_col].notnull())

        count = int(mask.sum())
        if count > 0:
            for idx in merged[mask].index:
                filled_cells.append({'row': idx + 2, 'col': col})

            if merged[col].dtype != merged[wni_col].dtype:
                merged[col] = merged[col].astype(object)

            merged.loc[mask, col] = merged.loc[mask, wni_col].values
            filled_total += count
            print(f"    OK  {col:<22} -> {count} cells filled")
        else:
            print(f"    --  {col:<22} -> nothing to fill")

    print(f"\n  TOTAL: {filled_total} cells filled")

    # 7. Cleanup and Save
    drop_cols = [c for c in merged.columns if c.endswith('_wni') or c == '_key']
    output_df = merged.drop(columns=drop_cols)

    print(f"\n  Saving to: {OUTPUT_FILENAME}...")
    output_df.to_excel(OUTPUT_FILENAME, index=False, sheet_name='Analysis Data')

    # 8. Apply Visual Formatting (Green Fill)
    print("  Applying green formatting to updated cells...")
    G_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    G_FONT = Font(color="276221", name="Arial", size=10)

    try:
        wb = load_workbook(OUTPUT_FILENAME)
        ws = wb.active
        col_map = {name: i+1 for i, name in enumerate(output_df.columns)}

        for item in filled_cells:
            c_idx = col_map.get(item['col'])
            if c_idx:
                cell = ws.cell(row=item['row'], column=c_idx)
                cell.fill = G_FILL
                cell.font = G_FONT
        wb.save(OUTPUT_FILENAME)
    except Exception as e:
        print(f"  Warning: Could not apply formatting: {e}")

    print("\n" + "="*60)
    print("  SUCCESS!")
    print(f"  Output File: {OUTPUT_FILENAME}")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()