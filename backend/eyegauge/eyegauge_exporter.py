"""
eyegauge_exporter.py
====================
Excel exporter for Eyegauge — reads ONLY from analysis_data table
where source_id = 'Eyegauge_API'.

MODE 1 — One-time full download (run this FIRST, only once):
    python eyegauge_exporter.py --mode download

MODE 2 — Append new rows (run after every pipeline execution):
    python eyegauge_exporter.py --mode append
    OR called automatically at the end of EyegaugePipeline.run()
"""

import argparse
import os
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

from ..config import config

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
# --- LOCAL (original hardcoded path — uncomment to use) ---
# EXPORT_DIR  = Path(r"C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\data\eyegauge\historical")
# --- VM / CROSS-PLATFORM (env var, default <project_root>/data/eyegauge/historical) ---
EXPORT_DIR  = Path(os.getenv("EYEGAUGE_EXPORT_DIR", str(config.ROOT_DIR / "data" / "eyegauge" / "historical")))
EXPORT_FILE = EXPORT_DIR / "eyegauge_analysis_data.xlsx"
SHEET_NAME  = "Analysis Data"
SOURCE_ID   = "Eyegauge_API"

# NOTE: NULL values from the DB are written as blank cells in Excel (pandas default).
# No replacement is applied — blank in DB = blank in Excel.


class EyegaugeExcelExporter:

    def __init__(self, engine: Engine):
        self.engine = engine

    # =========================================================================
    # MODE 1 — One-time full download
    # =========================================================================

    def download(self) -> Path:
        """
        Pull ALL rows from analysis_data where source_id = 'Eyegauge_API'
        and write to a brand-new Excel file.
        Run this ONCE before starting the pipeline.
        """
        logger.info("=== MODE: Initial Download ===")
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

        df = self._fetch_from_db()

        if df.empty:
            logger.warning("No data found in analysis_data for source_id = 'Eyegauge_API'.")
            return EXPORT_FILE

        logger.info(f"Fetched {len(df)} rows from analysis_data.")

        with pd.ExcelWriter(EXPORT_FILE, engine="openpyxl") as writer:
            self._write_sheet(writer, SHEET_NAME, df)

        logger.info(f"✓ Download complete — {len(df)} rows saved → {EXPORT_FILE}")
        return EXPORT_FILE

    # =========================================================================
    # MODE 2 — Append new rows after each pipeline run
    # =========================================================================

    def append_new(self) -> Path:
        """
        Read only NEW rows from analysis_data (not already in Excel)
        and append them below the existing data.
        Called automatically at the end of EyegaugePipeline.run().
        """
        logger.info("=== MODE: Append New Rows ===")

        if not EXPORT_FILE.exists():
            logger.warning("Excel file not found. Running full download first ...")
            return self.download()

        df_db = self._fetch_from_db()

        if df_db.empty:
            logger.info("No data in DB to append.")
            return EXPORT_FILE

        # Get IDs already present in the Excel
        existing_ids = self._get_existing_ids()
        df_new = df_db[~df_db["id"].astype(str).isin(existing_ids)].reset_index(drop=True)

        if df_new.empty:
            logger.info("No new rows to append. Excel is already up to date.")
            return EXPORT_FILE

        logger.info(f"Appending {len(df_new)} new rows to '{SHEET_NAME}' sheet ...")

        wb = load_workbook(EXPORT_FILE)
        self._append_to_sheet(wb, SHEET_NAME, df_new)
        wb.save(EXPORT_FILE)

        logger.info(f"✓ Append complete — {len(df_new)} new rows added → {EXPORT_FILE}")
        return EXPORT_FILE

    # =========================================================================
    # Private helpers
    # =========================================================================

    def _fetch_from_db(self) -> pd.DataFrame:
        """Fetch all rows from analysis_data where source_id = 'Eyegauge_API'."""
        try:
            with self.engine.connect() as conn:
                df = pd.read_sql(
                    text(
                        "SELECT * FROM analysis_data "
                        "WHERE source_id = :sid "
                        'ORDER BY "Date", "Time_UTC"'
                    ),
                    conn,
                    params={"sid": SOURCE_ID},
                )
            logger.info(f"DB query returned {len(df)} rows.")
            return df
        except Exception as e:
            logger.error(f"Failed to fetch from analysis_data: {e}")
            return pd.DataFrame()

    def _get_existing_ids(self) -> set:
        """Read id column from existing Excel sheet to find already-exported rows."""
        try:
            df_existing = pd.read_excel(EXPORT_FILE, sheet_name=SHEET_NAME, usecols=["id"])
            ids = set(df_existing["id"].dropna().astype(str).tolist())
            logger.info(f"Excel already has {len(ids)} rows.")
            return ids
        except Exception as e:
            logger.warning(f"Could not read existing Excel IDs: {e}")
            return set()

    def _write_sheet(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
        """Write a DataFrame to a sheet; skip if empty."""
        if df.empty:
            logger.warning(f"  ⚠ Sheet '{sheet_name}' — no data, skipped.")
            return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        logger.info(f"  ✓ Sheet '{sheet_name}' — {len(df)} rows written")

    def _append_to_sheet(self, wb, sheet_name: str, df: pd.DataFrame):
        """Append rows to an existing sheet in an openpyxl workbook."""
        if df.empty:
            return

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(list(df.columns))
        else:
            ws = wb[sheet_name]

        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))


# =============================================================================
# CLI entry point
# =============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Eyegauge Excel Exporter")
    parser.add_argument(
        "--mode",
        choices=["download", "append"],
        required=True,
        help=(
            "download = one-time full export from analysis_data to Excel | "
            "append   = add only new rows after pipeline run"
        ),
    )
    args = parser.parse_args()

    # eyegauge_exporter.py  →  backend/eyegauge/eyegauge_exporter.py
    # config.py             →  backend/config.py
    current_dir  = Path(__file__).resolve().parent   # backend/eyegauge
    backend_dir  = current_dir.parent                # backend
    project_root = backend_dir.parent                # Data_ingestion_pipeline

    for p in (backend_dir, project_root):
        if str(p) not in sys.path:
            sys.path.insert(0, str(p))

    try:
        from config import config
    except ModuleNotFoundError:
        try:
            from backend.config import config
        except Exception as e:
            logger.error(f"Could not import config: {e}")
            sys.exit(1)

    try:
        engine = create_engine(config.DATABASE_URL)
    except Exception as e:
        logger.error(f"Could not connect to DB: {e}")
        sys.exit(1)

    exporter = EyegaugeExcelExporter(engine)

    if args.mode == "download":
        exporter.download()
    elif args.mode == "append":
        exporter.append_new()