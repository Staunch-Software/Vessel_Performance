# ===========================================================================
# mariapps_exporter.py
#
# Exports analysis_data table from PostgreSQL → formatted Excel file.
# Saved to: C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\data\mariapps\historical
#
# - First run  : pulls ALL rows from DB → creates fresh Excel
# - Daily runs : pulls ALL rows from DB → appends only new rows below existing
#
# Called automatically at the end of mariapps_pipeline.py
# Can also be run standalone: python mariapps_exporter.py
# ===========================================================================

import os
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

from ..config import config

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
# --- LOCAL (original hardcoded URL — uncomment to use) ---
# DATABASE_URL = "postgresql://postgres:root@localhost:5432/noon_reports_db"
# --- VM / CROSS-PLATFORM (reads from .env via config) ---
DATABASE_URL = config.DATABASE_URL

# --- LOCAL (original hardcoded export dir — uncomment to use) ---
# EXPORT_DIR  = Path(r"C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\data\mariapps\historical")
# --- VM / CROSS-PLATFORM (env var, default <project_root>/data/mariapps/historical) ---
EXPORT_DIR  = Path(os.getenv("MARIAPPS_EXPORT_DIR", str(config.ROOT_DIR / "data" / "mariapps" / "historical")))
EXCEL_FILE  = "mariapps_historical_analysis.xlsx"
SHEET_NAME  = "Analysis Data"

# ---------------------------------------------------------------------------
# STYLING CONSTANTS
# ---------------------------------------------------------------------------
HEADER_BG  = "1F3864"
HEADER_FG  = "FFFFFF"
ALT_ROW_BG = "EBF0FA"
BORDER_COL = "BFBFBF"


def _thin_border():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _format_sheet(ws):
    """Apply header + alternating row formatting to a worksheet."""
    border     = _thin_border()
    total_cols = ws.max_column
    total_rows = ws.max_row

    # Header row
    for col_idx in range(1, total_cols + 1):
        cell            = ws.cell(row=1, column=col_idx)
        cell.font       = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill       = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = border
    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx in range(2, total_rows + 1):
        bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
        for col_idx in range(1, total_cols + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = border

    # Auto-fit column widths (capped at 40)
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len    = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(total_rows + 1, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb, combined_df, new_rows_count):
    """Create or overwrite the Summary sheet."""
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws       = wb.create_sheet("Summary")
    border   = _thin_border()
    hdr_fill = PatternFill("solid", start_color=HEADER_BG)
    hdr_font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)

    summary_rows = [
        ("Last Export Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source Table",          "analysis_data"),
        ("Database",              "noon_reports_db"),
        ("Total Records",         len(combined_df)),
        ("Total Columns",         len(combined_df.columns)),
        ("Earliest Log Date",     str(pd.to_datetime(combined_df["Date"], errors="coerce").min()) if "Date" in combined_df.columns else "N/A"),
        ("Latest Log Date",       str(pd.to_datetime(combined_df["Date"], errors="coerce").max()) if "Date" in combined_df.columns else "N/A"),
        ("Vessels Included",      combined_df["vessel_imo"].nunique() if "vessel_imo" in combined_df.columns else "N/A"),
        ("New Rows Added Today",  new_rows_count),
    ]

    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        lc           = ws.cell(row=r_idx, column=1, value=label)
        lc.font      = hdr_font
        lc.fill      = hdr_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = border

        vc           = ws.cell(row=r_idx, column=2, value=value)
        vc.font      = Font(name="Arial", size=10)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border    = border

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35

    # Keep Summary as first sheet
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames) - 1))


# ---------------------------------------------------------------------------
# MAIN EXPORT FUNCTION
# ---------------------------------------------------------------------------

def export_analysis_data(from_date: str = None, to_date: str = None) -> Path:
    """
    Pulls ALL rows from analysis_data → appends only new rows to the master Excel.

    Parameters (accepted for pipeline compatibility but not used for filtering)
    ----------
    from_date : str, optional  e.g. '2026-05-04'
    to_date   : str, optional  e.g. '2026-05-18'

    Returns
    -------
    Path to the saved Excel file.
    """
    log.info("=" * 60)
    log.info("  MariApps Excel Export — Starting")
    log.info("=" * 60)

    # --- Pull EVERY row from DB (no date filter — we need full history) ---
    engine = create_engine(DATABASE_URL)
    # TO:
    # TO:
    query  = text('''
        SELECT a.*, v.vessel_name 
        FROM analysis_data a
        LEFT JOIN vessels v ON a.vessel_imo = v.imo_number
        WHERE a.source_id = 'mari_apps' 
        ORDER BY a."Date" ASC, a.vessel_imo ASC
    ''')

    log.info("[DB]     Querying ALL rows from analysis_data ...")
    with engine.connect() as conn:
        db_df = pd.read_sql(query, conn)

    if db_df.empty:
        log.warning("[EXPORT] No data found in analysis_data. Excel file not created.")
        return None

    log.info(f"[DB]     Fetched {len(db_df)} rows x {len(db_df.columns)} columns from DB.")

    # --- Prepare output directory and file path ---
   # TO:
    # --- Prepare output directory ---
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    # --- Get unique vessel names from DB ---
    vessel_col = "vessel_imo"
    # Try to get vessel_name if available, else use imo
    if "vessel_name" in db_df.columns:
        vessel_col = "vessel_name"

    vessels = db_df[vessel_col].dropna().unique()
    log.info(f"[EXPORT] Found {len(vessels)} vessel(s) to export.")

    saved_paths = []

    for vessel in vessels:
        vessel_df  = db_df[db_df[vessel_col] == vessel].copy()
        safe_name  = str(vessel).strip()
        filename   = f"Mariapps_{safe_name}.xlsx"
        save_path  = EXPORT_DIR / filename

        log.info(f"[VESSEL] Processing: {safe_name} ({len(vessel_df)} rows)")

        # --- Append logic per vessel ---
        if save_path.exists():
            existing_df = pd.read_excel(save_path, sheet_name=SHEET_NAME, dtype=str)

            if "id" in existing_df.columns and "id" in vessel_df.columns:
                existing_ids   = set(existing_df["id"].dropna().astype(str))
                new_df         = vessel_df[~vessel_df["id"].astype(str).isin(existing_ids)]
            else:
                merge_cols     = [c for c in ["Date", "vessel_imo"] if c in existing_df.columns and c in vessel_df.columns]
                existing_keys  = set(existing_df[merge_cols].astype(str).apply(tuple, axis=1))
                new_df         = vessel_df[~vessel_df[merge_cols].astype(str).apply(tuple, axis=1).isin(existing_keys)]

            new_rows_count = len(new_df)
            log.info(f"[VESSEL] {new_rows_count} new row(s) to append for {safe_name}.")

            if new_rows_count == 0:
                log.info(f"[VESSEL] {safe_name} — already up to date.")
                wb = load_workbook(save_path)
                _write_summary_sheet(wb, existing_df, 0)
                wb.save(save_path)
                saved_paths.append(save_path)
                continue

            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

        else:
            log.info(f"[VESSEL] No existing file — creating fresh for {safe_name}.")
            combined_df    = vessel_df
            new_rows_count = len(vessel_df)

        # --- Write Excel ---
        combined_df.to_excel(save_path, index=False, sheet_name=SHEET_NAME, engine="openpyxl")

        wb = load_workbook(save_path)
        ws = wb[SHEET_NAME]
        _format_sheet(ws)
        _write_summary_sheet(wb, combined_df, new_rows_count)
        wb.save(save_path)

        log.info(f"[VESSEL] Saved: {filename} | New: {new_rows_count} | Total: {len(combined_df)}")
        saved_paths.append(save_path)

    log.info("=" * 60)
    log.info(f"[EXPORT] All vessels exported successfully!")
    log.info(f"[EXPORT]    Vessels   : {len(saved_paths)}")
    log.info(f"[EXPORT]    Directory : {EXPORT_DIR}")
    log.info("=" * 60)

    return saved_paths[0] if saved_paths else None


# ---------------------------------------------------------------------------
# STANDALONE ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    from_dt = sys.argv[1] if len(sys.argv) > 1 else None
    to_dt   = sys.argv[2] if len(sys.argv) > 2 else None

    export_analysis_data(from_date=from_dt, to_date=to_dt)