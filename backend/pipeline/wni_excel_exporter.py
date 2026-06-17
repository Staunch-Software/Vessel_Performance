import os
import datetime
import hashlib
import logging
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..database import SessionLocal
from ..models import Vessel, RawNoonReport, NoonReportData, AnalysisData, DataQualityLog, VesselParticulars
from .mapping import map_row, map_analysis_row, ANALYSIS_DATA_COLUMNS
from ..config import config

# --- LOCAL (original hardcoded path — uncomment to use) ---
# EXCEL_BASE_PATH = r"C:\Users\visha\Downloads\ozellar\Data_ingestion_pipeline\data\wni\historical"
# --- VM / CROSS-PLATFORM (env var, default <project_root>/data/wni/historical) ---
EXCEL_BASE_PATH = os.getenv("WNI_EXPORT_DIR", str(config.ROOT_DIR / "data" / "wni" / "historical"))


def get_excel_path(vessel_name: str) -> str:
    return os.path.join(EXCEL_BASE_PATH, f"WNI_{vessel_name.strip().upper()}.xlsx")


def _write_header(ws):
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="003366")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, col_name in enumerate(ANALYSIS_DATA_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30
    for col_idx, col_name in enumerate(ANALYSIS_DATA_COLUMNS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(col_name) + 2, 14)


def _append_rows(ws, rows: list):
    next_row = ws.max_row + 1
    data_font = Font(name="Arial", size=9)
    data_align = Alignment(horizontal="left", vertical="center")
    for row_dict in rows:
        for col_idx, col_name in enumerate(ANALYSIS_DATA_COLUMNS, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=row_dict.get(col_name))
            cell.font = data_font
            cell.alignment = data_align
        next_row += 1


def export_historical_to_excel():
    """First-run: pulls all AnalysisData where source_id = wni, creates one Excel file per vessel."""
    db = SessionLocal()
    try:
        records = db.query(AnalysisData).filter(AnalysisData.source_id == "wni").order_by(AnalysisData.Date.asc(), AnalysisData.Time_UTC.asc()).all()
        if not records:
            logging.warning("No WNI records found in analysis_data table.")
            return

        grouped = defaultdict(list)
        for r in records:
            grouped[r.vessel_imo].append(r)

        os.makedirs(EXCEL_BASE_PATH, exist_ok=True)

        for vessel_imo, vessel_records in grouped.items():
            vessel = db.query(Vessel).filter(Vessel.imo_number == vessel_imo).first()
            vessel_name = vessel.vessel_name if vessel else str(vessel_imo)
            rows = [{col: getattr(r, col, None) for col in ANALYSIS_DATA_COLUMNS} for r in vessel_records]
            excel_path = get_excel_path(vessel_name)
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis Data"
            ws.freeze_panes = "A2"
            _write_header(ws)
            _append_rows(ws, rows)
            wb.save(excel_path)
            logging.info(f"Exported: {excel_path} ({len(rows)} rows)")

    except Exception as e:
        logging.error(f"Historical export failed: {e}")
    finally:
        db.close()


def append_to_excel(vessel_name: str, analysis_row: dict):
    """Daily run: appends one new row to the existing Excel file.
    If the Excel does not exist yet, runs a full historical export first
    (which already includes the current row since it is committed to DB),
    then returns — no double-append.
    """
    excel_path = get_excel_path(vessel_name)

    if not os.path.exists(excel_path):
        logging.info(f"Excel not found for '{vessel_name}'. Running historical export to create it.")
        export_historical_to_excel()
        # Historical export already contains this row (it was committed before
        # this function was called), so we are done.
        return

    try:
        wb = load_workbook(excel_path)
        ws = wb["Analysis Data"]

        # Guard against duplicate append: check if a row with the same Date
        # and Time_UTC already exists in the sheet.
        new_date = analysis_row.get("Date")
        new_time = analysis_row.get("Time_UTC")
        date_col  = ANALYSIS_DATA_COLUMNS.index("Date") + 1
        time_col  = ANALYSIS_DATA_COLUMNS.index("Time_UTC") + 1

        for row in ws.iter_rows(min_row=2, max_col=max(date_col, time_col), values_only=True):
            if str(row[date_col - 1]) == str(new_date) and str(row[time_col - 1]) == str(new_time):
                logging.warning(
                    f"Skipping append for '{vessel_name}': row {new_date} {new_time} already exists in Excel."
                )
                wb.close()
                return

        _append_rows(ws, [analysis_row])
        wb.save(excel_path)
        logging.info(f"Appended 1 row to {excel_path} | date={new_date} time={new_time}")

    except Exception as e:
        logging.error(f"Append failed for '{vessel_name}': {e}")


def clean_dict(d):
    return {k: (None if pd.isna(v) else v) for k, v in d.items()}


def get_audit_label(source_id):
    now = datetime.datetime.now()
    source_tag = str(source_id).upper()
    month = now.strftime('%b').upper()
    week_of_month = (now.day - 1) // 7 + 1
    year = now.year
    return f"{source_tag}-{month}-W{week_of_month}-{year}"


def save_to_db(vessel_name_raw, raw_data_dict, file_name):
    db = SessionLocal()
    try:
        clean_name = vessel_name_raw.split("/")[0].strip()
        vessel = db.query(Vessel).filter(Vessel.vessel_name.ilike(clean_name)).first()
        if not vessel:
            logging.error(f"Vessel '{clean_name}' not found in DB.")
            return "error"

        cleaned_json = clean_dict(raw_data_dict)

        raw_date_str = str(cleaned_json.get("Date", "")).strip()
        event_type = str(cleaned_json.get("Event Type", "")).strip().upper()
        fp_str = f"{vessel.imo_number}|{raw_date_str}|{event_type}"
        fingerprint = hashlib.sha256(fp_str.encode()).hexdigest()

        is_duplicate = db.query(RawNoonReport).filter(RawNoonReport.fingerprint == fingerprint).first() is not None

        raw_rec = RawNoonReport(
            vessel_imo=vessel.imo_number,
            source_id="wni",
            raw_json=cleaned_json,
            file_name=file_name,
            fingerprint=fingerprint
        )
        db.add(raw_rec)
        db.flush()

        if is_duplicate:
            db.add(DataQualityLog(
                raw_report_id=raw_rec.id,
                source_id="wni",
                vessel_name=vessel.vessel_name,
                vessel_imo=vessel.imo_number,
                issue_type="DUPLICATE_REPORT",
                event_type=cleaned_json.get('Event Type'),
                report_date=pd.to_datetime(cleaned_json.get('Date'), errors='coerce'),
                audit_period=get_audit_label("wni")
            ))
            db.commit()
            return "duplicate"

        noon_data = map_row(cleaned_json).to_dict()

        print(f"DEBUG: log_date_utc value: {noon_data.get('log_date_utc')}")
        print(f"DEBUG: log_type value: {noon_data.get('log_type')}")

        if not noon_data.get('log_date_utc') or not noon_data.get('log_type'):
            logging.error(f"Mandatory fields missing for {vessel.vessel_name}")
            db.rollback()
            return "error"

        db.add(NoonReportData(
            **clean_dict(noon_data),
            vessel_imo=vessel.imo_number,
            source_id="wni",
            raw_report_id=raw_rec.id
        ))

        specs = db.query(VesselParticulars).filter(VesselParticulars.vessel_imo == vessel.imo_number).first()
        if not specs:
            logging.warning(f"No VesselParticulars found for IMO {vessel.imo_number}. Using defaults.")

        analysis_data = map_analysis_row(cleaned_json, specs=specs)

        db.add(AnalysisData(
            **clean_dict(analysis_data),
            vessel_imo=vessel.imo_number,
            source_id="wni",
            raw_report_id=raw_rec.id
        ))

        db.commit()

        # Append new row to Excel after successful DB save
        append_to_excel(vessel.vessel_name, analysis_data)

        return "success"

    except Exception as e:
        db.rollback()
        logging.error(f"DB Error for {vessel_name_raw}: {e}")
        return "error"

    finally:
        db.close()


if __name__ == "__main__":
    export_historical_to_excel()